# secop_extract.py
from __future__ import annotations

import re
import time
import random
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import openpyxl
from bs4 import BeautifulSoup

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


# -----------------------------
# Configuracion
# -----------------------------
SECOP_BASE_URL = "https://www.contratos.gov.co/consultas/detalleProceso.do?numConstancia="

# Constancia tipo: 25-1-241304, 25-15-14542595, etc.
# Nota: en la practica SECOP I usa yy-m-xxxxxx o yy-mm-xxxxxx; toleramos 1-2 digitos en el bloque "xx".
CONSTANCIA_RE = re.compile(r"^(?P<yy>\d{2})-(?P<xx>\d{1,2})-(?P<num>\d{4,12})$")

# Soporta guiones Unicode comunes (en dash, em dash, etc.)
DASHES_RE = re.compile(r"[\u2010\u2011\u2012\u2013\u2014\u2212]")

# Señales tipicas de bloqueo anti-DDoS del sitio
BLOCK_MARKERS = [
    "access blocked",
    "acceso bloqueado",
    "possible ddos",
    "denegacion",
    "hic",
    "incident id",
]


# -----------------------------
# Errores
# -----------------------------
class SecopExtractionError(Exception):
    pass


def normalize_constancia(constancia: str) -> str:
    s = (constancia or "").strip()
    s = DASHES_RE.sub("-", s)
    s = re.sub(r"\s+", "", s)
    return s


def validate_constancia(constancia: str) -> str:
    c = normalize_constancia(constancia)
    if not CONSTANCIA_RE.match(c):
        raise SecopExtractionError(
            f"Constancia invalida: '{constancia}'. Formato esperado: YY-XX-NNNN (ej.: 25-1-241304, 25-15-14542595)"
        )
    return c


def build_url(constancia: str) -> str:
    return f"{SECOP_BASE_URL}{constancia}"


def _safe_text(node: Any) -> str:
    if node is None:
        return ""
    return node.get_text(" ", strip=True) if hasattr(node, "get_text") else str(node).strip()


def _norm_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    s = s.replace(":", "")
    return s


def _extract_kv_from_table(table) -> List[Tuple[str, str]]:
    """
    Extrae pares (label, value) de una tabla tipica de 2 columnas.
    Respeta textareas (objetos largos).
    """
    pairs: List[Tuple[str, str]] = []
    if not table:
        return pairs
    for tr in table.find_all("tr"):
        tds = tr.find_all("td", recursive=False)
        if len(tds) < 2:
            continue
        label = _safe_text(tds[0]).strip()
        value_td = tds[1]
        if not label:
            continue
        textarea = value_td.find("textarea")
        if textarea:
            value = textarea.get_text("\n", strip=True)
        else:
            value = _safe_text(value_td)
        if value is None:
            value = ""
        pairs.append((label, value))
    return pairs


def _find_section_header_td(soup: BeautifulSoup, header_text: str):
    target = _norm_key(header_text)
    for td in soup.find_all("td"):
        if "tttablas" in (td.get("class") or []):
            key = _norm_key(_safe_text(td))
            if key == target or (target and target in key) or (key and key in target):
                return td
    return None


def _parse_section_kv(soup: BeautifulSoup, header_text: str) -> List[Tuple[str, str]]:
    """
    Dado un encabezado de seccion (td.tttablas), encuentra la primera tabla util posterior y la parsea como KV.
    """
    header_td = _find_section_header_td(soup, header_text)
    if not header_td:
        return []
    header_tr = header_td.find_parent("tr")
    if not header_tr:
        return []
    # Buscar tablas siguientes; tomar la primera que tenga al menos 2 filas KV plausibles
    for table in header_tr.find_all_next("table", limit=12):
        pairs = _extract_kv_from_table(table)
        # Heuristica: al menos 3 pares y labels con algo de contenido
        if len(pairs) >= 3:
            return pairs
    return []



def _is_nonempty(v: str) -> bool:
    return bool(v and str(v).strip() and str(v).strip().lower() not in {"nan", "none", "null", "-"})


def _merge_maps_keep_first(*maps: dict) -> dict:
    """Merge maps preserving the first non-empty value for each key."""
    out = {}
    for mp in maps:
        if not mp:
            continue
        for k, v in mp.items():
            if k not in out or not _is_nonempty(out.get(k, "")):
                if _is_nonempty(v):
                    out[k] = v
    return out




def _norm_text(s: str) -> str:
    """Normaliza texto para comparaciones tolerantes (sin tildes, sin puntuacion, espacios colapsados)."""
    if s is None:
        return ""
    s = str(s)
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_row_value_by_label(soup, label_substr: str) -> str:
    """Busca valor (celda derecha) para una fila cuyo rotulo (celda izquierda) coincide de forma tolerante."""
    target = _norm_text(label_substr)
    if not target:
        return ""
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            left = _safe_text(cells[0])
            if not left:
                continue
            if target in _norm_text(left):
                return _safe_text(cells[1]).strip()
    return ""

def _find_rp_code(soup) -> str:
    """Extrae el codigo RP/CRP desde la tabla con encabezados 'Codigo|Fecha|Valor' de forma tolerante.
    No depende del titulo de seccion (SECOP varia el encabezado).
    """
    # 1) Tabla por estructura (encabezados)
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        header_cells = rows[0].find_all(["th", "td"])
        headers = [_norm_text(_safe_text(c)) for c in header_cells]
        if not headers:
            continue
        if "codigo" in headers and "valor" in headers:
            code_idx = headers.index("codigo")
            for r in rows[1:]:
                data_cells = r.find_all(["th", "td"])
                if len(data_cells) <= code_idx:
                    continue
                raw = _safe_text(data_cells[code_idx])
                raw_digits = _extract_digits(raw)
                if raw_digits and len(raw_digits) >= 6:
                    return raw_digits
    # 2) Fallback regex en texto completo
    text = soup.get_text(" ", strip=True)
    for pat in (
        r"\bRP\b\s*(?:No\.|Nro\.|#|:)?\s*(\d{6,})",
        r"\bCRP\b\s*(?:No\.|Nro\.|#|:)?\s*(\d{6,})",
    ):
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return m.group(1)
    return ""

def _parse_all_kv(soup):
    """Parsea pares etiqueta/valor de manera tolerante recorriendo toda la pagina.

    - Filas con 2 celdas (th/td o td/td)
    - Ignora tablas de documentos/hitos cuando no tienen estructura KV
    """
    pairs = []
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"], recursive=False)
            if len(cells) != 2:
                # algunos layouts tienen td anidados; intentar modo recursivo si no hay celdas directas
                cells = tr.find_all(["th", "td"])[:2]
            if len(cells) != 2:
                continue
            k = _safe_text(cells[0])
            v = _safe_text(cells[1])
            if _is_nonempty(k) and _is_nonempty(v):
                pairs.append((k, v))
    return pairs


def _parse_section_table(soup: BeautifulSoup, header_text: str) -> Optional[List[List[str]]]:
    """
    Parsea una tabla con encabezados (th/td) posterior a un header.
    Retorna matriz (filas) incluyendo encabezado como primera fila.
    """
    header_td = _find_section_header_td(soup, header_text)
    if not header_td:
        return None
    header_tr = header_td.find_parent("tr")
    if not header_tr:
        return None
    for table in header_tr.find_all_next("table", limit=15):
        rows = []
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"], recursive=False)
            if not cells:
                continue
            rows.append([_safe_text(c) for c in cells])
        # Heuristica: tabla con encabezado y 1+ filas
        if len(rows) >= 2 and len(rows[0]) >= 2:
            return rows
    return None


def _kv_to_map(pairs: List[Tuple[str, str]]) -> Dict[str, str]:
    m: Dict[str, str] = {}
    for k, v in pairs:
        nk = _norm_key(k)
        if nk and nk not in m:  # primera ocurrencia gana
            m[nk] = (v or "").strip()
    return m


def _get_first(m: Dict[str, str], keys: List[str]) -> str:
    for k in keys:
        nk = _norm_key(k)
        # match exact
        if nk in m and m[nk]:
            return m[nk]
    # fallback: contains tokens
    for k in keys:
        nk = _norm_key(k)
        for mk, mv in m.items():
            if mv and nk in mk:
                return mv
    return ""


def _parse_numero_proceso_informativo(soup: BeautifulSoup) -> str:
    for line in soup.get_text("\n", strip=True).splitlines():
        if "detalle del proceso numero" in _norm_text(line):
            parts = line.split(":", 1)
            if len(parts) == 2:
                return parts[1].strip()
    return ""


def _parse_fuente_financiacion(soup: BeautifulSoup) -> str:
    # Preferir tabla "Fuentes de Financiacion"
    rows = _parse_section_table(soup, "Fuentes de Financiacion")
    if rows and len(rows) >= 2:
        header = [_norm_key(h) for h in rows[0]]
        # buscamos columna "fuente"
        try:
            idx_fuente = header.index("fuente")
        except ValueError:
            idx_fuente = None
        fuentes = []
        for r in rows[1:]:
            if idx_fuente is not None and idx_fuente < len(r):
                val = (r[idx_fuente] or "").strip()
                if val:
                    fuentes.append(val)
        if fuentes:
            # dedupe manteniendo orden
            out=[]
            for f in fuentes:
                if f not in out:
                    out.append(f)
            return "; ".join(out)
    return ""


def _parse_rp_table(soup: BeautifulSoup) -> Dict[str, str]:
    """
    Extrae Codigo RP/CRP, Fecha y Valor desde "Registro Presupuestal del Compromiso (RP)".
    Retorna dict con claves: codigo_rp, fecha_rp, valor_rp
    """
    out = {"codigo_rp": "", "fecha_rp": "", "valor_rp": ""}
    # intentar varios encabezados posibles (SECOP I varia)
    rows = None
    for h in [
        "Registro Presupuestal del Compromiso (R.P.)",
        "Registro Presupuestal del Compromiso (RP)",
        "Registro Presupuestal del Compromiso",
        "Registro Presupuestal (RP)",
        "Registro Presupuestal",
        "Registro Presupuestal del Compromiso - RP",
    ]:
        rows = _parse_section_table(soup, h)
        if rows:
            break
    if not rows:
        return out
    header = [_norm_key(h) for h in rows[0]]
    # indices posibles
    def idx_of(*cands):
        for c in cands:
            nc=_norm_key(c)
            if nc in header:
                return header.index(nc)
        return None
    i_codigo = idx_of("Codigo", "Codigo")
    i_fecha = idx_of("Fecha")
    i_valor = idx_of("Valor")
    # Tomar primera fila de datos valida
    for r in rows[1:]:
        if not r or len(r) < 2:
            continue
        if i_codigo is not None and i_codigo < len(r) and not out["codigo_rp"]:
            out["codigo_rp"] = (r[i_codigo] or "").strip()
        if i_fecha is not None and i_fecha < len(r) and not out["fecha_rp"]:
            out["fecha_rp"] = (r[i_fecha] or "").strip()
        if i_valor is not None and i_valor < len(r) and not out["valor_rp"]:
            out["valor_rp"] = (r[i_valor] or "").strip()
        if out["codigo_rp"] or out["fecha_rp"] or out["valor_rp"]:
            break
    return out


def _extract_cdp(soup: BeautifulSoup) -> str:
    """Extrae el certificado de disponibilidad presupuestal (CDP) de forma tolerante."""
    raw = _find_row_value_by_label(soup, "Numero del respaldo presupuestal")
    if raw:
        token = _pick_numeric_token(raw)
        if token:
            return token

    rows = _parse_section_table(soup, "Respaldos Presupuestales Asociados al Proceso")
    if rows:
        header = [_norm_text(h) for h in rows[0]]
        idx_num = None
        for i, h in enumerate(header):
            if "numero" in h and "respaldo" in h:
                idx_num = i
        if idx_num is not None:
            candidates: List[str] = []
            for r in rows[1:]:
                if idx_num < len(r):
                    val = (r[idx_num] or "").strip()
                    if val:
                        token = _pick_numeric_token(val)
                        if not token:
                            continue
                        if len(token) == 10:
                            return token
                        candidates.append(token)
            if candidates:
                return max(candidates, key=len)

    raw = _find_row_value_by_label(soup, "Certificado de disponibilidad presupuestal")
    if not raw:
        raw = _find_row_value_by_label(soup, "CDP")
    if raw:
        token = _pick_numeric_token(raw)
        if token:
            return token

    text_all = soup.get_text(" ", strip=True)
    m = re.search(r"\bCDP\b\s*(?:No\.|Nro\.|#|:)?\s*([A-Za-z0-9\-/]+)", text_all, flags=re.IGNORECASE)
    if m:
        token = m.group(1).strip()
        picked = _pick_numeric_token(token)
        if picked:
            return picked
    m = re.search(r"respaldo presupuestal\s*(?:No\.|Nro\.|#|:)?\s*([A-Za-z0-9\-/]+)", text_all, flags=re.IGNORECASE)
    if m:
        token = m.group(1).strip()
        picked = _pick_numeric_token(token)
        if picked:
            return picked
    return ""


def fetch_detail_html(constancia: str, headless: bool = False, timeout_ms: int = 120_000) -> str:
    """
    Abre el detalle SECOP I en Playwright (visible por defecto para resolver reCAPTCHA) y retorna el HTML renderizado.
    """
    url = build_url(constancia)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1280, "height": 720})
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            # Espera razonable a tablas principales. Si hay reCAPTCHA, el usuario debe resolverlo y luego continuar.
            # Reintentamos esperar por un elemento comun en el detalle.
            page.wait_for_timeout(1500)
            try:
                page.wait_for_selector("td.tttablas", timeout=20_000)
            except PWTimeoutError:
                # No forzamos fallo inmediato; puede haber variaciones o el usuario aun resolviendo captcha.
                pass
            # Espera adicional corta para render
            page.wait_for_timeout(1200)
            html = page.content()
        finally:
            context.close()
            browser.close()
    return html


def _is_blocked_html(html: str) -> bool:
    text = (html or "").lower()
    return any(marker in text for marker in BLOCK_MARKERS)


def _fetch_detail_html_with_page(page, constancia: str, timeout_ms: int = 120_000) -> str:
    """
    Variante para reusar un page/contexto en lotes y evitar señales de automatizacion agresiva.
    """
    url = build_url(constancia)
    page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
    page.wait_for_timeout(1500)
    try:
        page.wait_for_selector("td.tttablas", timeout=20_000)
    except PWTimeoutError:
        pass
    page.wait_for_timeout(1200)
    html = page.content()
    if _is_blocked_html(html):
        raise SecopExtractionError(
            "Acceso bloqueado por el sitio (posible DDoS/WAF). Deteniendo el lote; esperar y/o contactar soporte."
        )
    return html



def _extract_digits(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    d = re.sub(r"[^0-9]", "", s)
    return d


def _pick_numeric_token(s: str) -> str:
    """Devuelve el token numerico mas probable en un string.

    Prioriza formatos tipo YYMMDD#### (10 digitos). Si no existe, devuelve
    el primer token razonable (4-9 digitos) o el mas largo disponible.
    """
    s = (s or "").strip()
    if not s:
        return ""
    tokens = re.findall(r"\d+", s)
    if not tokens:
        return ""
    for t in tokens:
        if len(t) == 10:
            return t
    for t in tokens:
        if 4 <= len(t) <= 9:
            return t
    return max(tokens, key=len)



def _clean_id(s: str) -> str:
    """Identificacion limpia (solo digitos)."""
    return _extract_digits(s)


def _extract_id_type(raw: str) -> str:
    """Extrae el tipo de identificacion (CC, NIT, CE, etc.) desde un string."""
    s = (raw or "").strip()
    if not s:
        return ""
    s_norm = _norm_text(s)
    tokens = s_norm.split()
    if "nit" in tokens:
        return "NIT"
    if "extranjeria" in tokens or "ce" in tokens:
        return "CE"
    if "pasaporte" in tokens or "pas" in tokens:
        return "PAS"
    if "pep" in tokens:
        return "PEP"
    if ("tarjeta" in tokens and "identidad" in tokens) or "ti" in tokens:
        return "TI"
    if "cedula" in tokens or "cc" in tokens:
        return "CC"
    m = re.match(r"\s*([A-Za-z]{1,5})\b", s)
    if m:
        token = m.group(1).upper()
        if token not in {"NO", "NRO"}:
            return token
    return ""


def _clean_bpim(raw: str) -> str:
    """Normaliza BPIN/BPIM a un unico codigo numerico sin leyendas.

    Objetivo: YYYY + consecutivo (p.ej. 20250000003856). El portal puede traer:
    - 'BPIM: 2025 00000003856'
    - 'Codigo BPIM Ano 2025 20250000003856'
    - Variaciones con puntos/espacios.
    """
    s = (raw or "").strip()
    if not s:
        return ""
    # extraer todos los grupos numericos
    nums = re.findall(r"\d+", s)
    if not nums:
        return ""

    # Si ya existe un token largo que empieza por 20xx, usar el mas largo.
    candidates = [n for n in nums if len(n) >= 12 and n.startswith("20")]
    if candidates:
        # elegir el mas largo; si empate, el primero
        candidates.sort(key=lambda x: (-len(x), nums.index(x)))
        return candidates[0]

    # Caso tipico separado: '2025' + '00000003856' -> concatenar
    year_tokens = [n for n in nums if len(n) == 4 and n.startswith("20")]
    if year_tokens:
        year = year_tokens[0]
        # elegir el token mas largo que NO sea el ano y que no empiece por 20
        others = [n for n in nums if n != year]
        if others:
            other = max(others, key=len)
            merged = year + other
            if len(merged) >= 12:
                return merged

    # Fallback: si hay un token largo sin '20', devolver el mas largo
    other_long = [n for n in nums if len(n) >= 10]
    if other_long:
        return max(other_long, key=len)

    return ""


def _clean_fuente_financiacion(raw: str) -> str:
    """Normaliza la fuente de financiacion a un catalogo corto y util.

    Ejemplos de salida:
    - 'Sistema General de Regalias'
    - 'Sistema General de Participaciones (SGP)'
    - 'Recursos propios'
    - 'Otros recursos'
    """
    s = (raw or "").strip()
    if not s:
        return ""
    n = _norm_key(s)

    # Prioridad por especificidad
    if "regalia" in n or "sgr" in n:
        return "Sistema General de Regalias"
    if "participacion" in n or re.search(r"\bsgp\b", n):
        return "Sistema General de Participaciones (SGP)"
    if "propio" in n:
        return "Recursos propios"
    if "otro" in n or "otros" in n:
        return "Otros recursos"

    # Si no clasifica, devolver el texto original recortado (sin encabezados comunes)
    # Quitar palabras de encabezado frecuentes
    for junk in ["fuente", "financiacion", "financiacion", "valor", "tipo"]:
        n = n.replace(junk, "").strip()
    # Si quedo vacio, devolver original
    return s if n else s


def _clean_money(s: str) -> str:
    """Convierte un valor monetario a entero (COP) truncando decimales de forma segura.

    Evita inflar el valor cuando la fuente trae ',00' o '.00' y se eliminan separadores.
    Soporta formatos tipicos:
      - es-CO: 608.603.520.000,00
      - en-US: 608,603,520,000.00
      - enteros: 608603520000
    Devuelve string de digitos (sin separadores)."""
    s = (s or "").strip()
    if not s:
        return ""
    # Dejar solo digitos y separadores de miles/decimales
    raw = re.sub(r"[^0-9,\.]", "", s)
    if not raw:
        return ""

    # Caso mixto con ambos separadores: decidir decimal por el ultimo separador
    if "," in raw and "." in raw:
        last_comma = raw.rfind(",")
        last_dot = raw.rfind(".")
        if last_comma > last_dot:
            # decimal = ',' ; miles = '.'
            int_part = raw.split(",", 1)[0]
            return re.sub(r"[^0-9]", "", int_part)
        else:
            # decimal = '.' ; miles = ','
            int_part = raw.split(".", 1)[0]
            return re.sub(r"[^0-9]", "", int_part)

    # Solo coma
    if "," in raw and "." not in raw:
        left, right = raw.split(",", 1)
        # si son 2 digitos al final, asumimos decimal
        if right.isdigit() and len(right) == 2:
            return re.sub(r"[^0-9]", "", left)
        # si no, es separador de miles
        return re.sub(r"[^0-9]", "", raw)

    # Solo punto
    if "." in raw and "," not in raw:
        left, right = raw.split(".", 1)
        if right.isdigit() and len(right) == 2:
            return re.sub(r"[^0-9]", "", left)
        return re.sub(r"[^0-9]", "", raw)

    # Solo digitos
    return re.sub(r"[^0-9]", "", raw)

def _extract_rp_code(raw: str) -> str:
    """Normaliza RP/CRP a un codigo simple (solo digitos)."""
    s = (raw or "").strip()
    if not s:
        return ""
    m = re.search(r"(\d{6,})", s)
    if m:
        return m.group(1)
    return re.sub(r"[^\d]", "", s)


def _extract_crp_code(soup: BeautifulSoup) -> str:
    """Extrae el codigo CRP usando tabla + fallback (elimina duplicacion)."""
    rp = _parse_rp_table(soup)
    crp_from_table = _extract_rp_code(rp.get("codigo_rp", ""))
    if not crp_from_table:
        crp_fallback = _find_rp_code(soup)
        return crp_fallback
    return crp_from_table



def _determine_tipo_proceso(tipo_gasto: str) -> str:
    # Normaliza para tolerar tildes y variaciones ("Inversion"/"Inversión")
    s = _norm_text(tipo_gasto)
    if "inversion" in s:
        return "Inversion"
    if "funcionamiento" in s:
        return "Funcionamiento"
    return ""


def _estado_validacion(record: Dict[str, str]) -> Tuple[str, str]:
    # Campos criticos para que el cuadro sea util
    critical = {
        "Modalidad de contratacion": record.get("Modalidad de contratacion", ""),
        "Objeto del contrato": record.get("Objeto del contrato", ""),
        "Valor del contrato (COP)": record.get("Valor del contrato (COP)", ""),
        "Razon social del proponente/contratista": record.get("Razon social del proponente/contratista", ""),
    }

    missing = [k for k, v in critical.items() if not (v or "").strip()]
    if not missing:
        return "Completo", ""
    if len(missing) == 1:
        return "Revision", f"Falta: {missing[0]}"
    return "Incompleto", "Faltan: " + "; ".join(missing)


def _load_template(template_path: Path):
    if not template_path.exists():
        raise SecopExtractionError(f"No se encontro la plantilla: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    if "Resultados_Extraccion" not in wb.sheetnames:
        raise SecopExtractionError("La plantilla no contiene la hoja 'Resultados_Extraccion'.")
    return wb


def _append_errors_sheet(wb, errors: List[Tuple[str, str]]) -> None:
    if not errors:
        return
    if "Errores" in wb.sheetnames:
        ws_err = wb["Errores"]
    else:
        ws_err = wb.create_sheet("Errores")
        ws_err.append(["numConstancia", "error"])
    for c, err in errors:
        ws_err.append([c, err])


def _find_next_row(ws) -> int:
    """
    Encuentra la siguiente fila vacia basada en columnas clave (evita columna con formula).
    """
    headers = [c.value for c in ws[1]]
    header_map = {_norm_key(str(h or "")): i + 1 for i, h in enumerate(headers)}
    col_a = header_map.get(_norm_key("Numero de proceso (informativo)"), 1)
    col_b = header_map.get(_norm_key("Numero de constancia"), 2)
    for r in range(2, ws.max_row + 2):
        a = ws.cell(row=r, column=col_a).value
        b = ws.cell(row=r, column=col_b).value
        if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == ""):
            return r
    return ws.max_row + 1


def _get_hyperlink_base(wb) -> str:
    base = ""
    if "Config" in wb.sheetnames:
        base = str(wb["Config"]["B1"].value or "").strip()
    if not base:
        base = SECOP_BASE_URL
    return base


def _write_record_row(ws, wb, headers, record: Dict[str, str], constancia_ok: str, row_idx: int) -> None:
    record_norm = {_norm_key(k): v for k, v in record.items()}
    base = _get_hyperlink_base(wb)
    for col_idx, h in enumerate(headers, start=1):
        if not h:
            continue
        h_norm = _norm_key(str(h))
        if h_norm == _norm_key("Abrir detalle"):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = "Abrir"
            cell.hyperlink = base + constancia_ok
            try:
                cell.style = "Hyperlink"
            except Exception:
                pass
            continue
        if h_norm in record_norm:
            ws.cell(row=row_idx, column=col_idx, value=record_norm[h_norm])


def _build_record_from_soup(soup: BeautifulSoup, constancia_ok: str) -> Dict[str, str]:
    # Extracciones dirigidas (sin depender de secciones): representante legal e RP
    rep_id_raw = _find_row_value_by_label(soup, "Identificacion del Representante Legal")

    # 0) Baseline KV tolerante (anti-regresion)
    baseline_pairs = _parse_all_kv(soup)
    baseline_map = _kv_to_map(baseline_pairs)

    # 1) General (KV por seccion) - si no se encuentra, se apoya en baseline_map
    general_pairs = _parse_section_kv(soup, "Informacion General del Proceso")
    general_map = _merge_maps_keep_first(_kv_to_map(general_pairs), baseline_map)

    # 2) Contrato (KV por seccion) - si no se encuentra, se apoya en baseline_map
    contrato_pairs = _parse_section_kv(soup, "Informacion del Contrato")
    contrato_map = _merge_maps_keep_first(_kv_to_map(contrato_pairs), baseline_map)

    # 3) Presupuestal (RP table + fallback KV)
    # Prioridad RP: tabla presupuestal de la seccion; fallback conservador a busqueda tolerante
    rp_code = _extract_crp_code(soup)
    cdp = _extract_cdp(soup)

    # Campo informativo "Numero de proceso"
    num_proceso_info = _parse_numero_proceso_informativo(soup)

    modalidad = _get_first(general_map, ["Tipo de Proceso", "Modalidad de Contratacion", "Modalidad"])
    estado_proc = _get_first(general_map, ["Estado del Proceso", "Estado del Contrato", "Estado"])

    fuente_fin = _parse_fuente_financiacion(soup)
    if not fuente_fin:
        fuente_fin = _get_first(general_map, ["Fuente de Financiacion", "Fuentes de Financiacion", "Fuente"])

    fuente_fin = _clean_fuente_financiacion(fuente_fin)

    # Contrato info
    num_contrato = _get_first(contrato_map, ["Numero del Contrato", "No. Contrato", "Contrato No", "Numero de Contrato"])
    objeto = _get_first(contrato_map, ["Objeto del Contrato", "Objeto"])
    valor = _get_first(contrato_map, ["Cuantia Definitiva del Contrato", "Cuantia del Contrato", "Valor del Contrato", "Cuantia", "Valor"])
    valor_num = _clean_money(valor)

    plazo = _get_first(contrato_map, ["Plazo de Ejecucion del Contrato", "Plazo de Ejecucion", "Plazo"])
    fecha_inicio = _get_first(contrato_map, ["Fecha de Inicio de Ejecucion del Contrato", "Fecha de Inicio", "Fecha inicio"])
    fecha_fin = _get_first(contrato_map, ["Fecha de Terminacion del Contrato", "Fecha de Terminacion", "Fecha fin", "Fecha terminacion"])

    razon_social = _get_first(contrato_map, ["Nombre o Razon Social del Contratista", "Nombre o Razon Social del Contratista", "Contratista", "Adjudicatario"])
    ident = _get_first(contrato_map, ["Identificacion del Contratista", "Identificacion del Contratista", "NIT del Contratista", "NIT", "Cedula", "Cedula", "Identificacion", "Identificacion"])
    if not ident:
        ident = _get_first(general_map, ["Identificacion", "Identificacion", "NIT", "Cedula", "Cedula"])
    rep_legal = _get_first(contrato_map, ["Nombre del Representante Legal del Contratista", "Representante Legal", "Representante"])
    if not rep_legal:
        rep_legal = _get_first(general_map, ["Representante Legal", "Representante"])

    rep_ident = _get_first(
        contrato_map,
        [
            "Identificacion del Representante Legal del Contratista",
            "Identificacion del Representante Legal",
            "Identificacion Representante Legal",
            "Cedula Representante",
            "Identificacion Representante",
        ],
    )
    if not rep_ident:
        rep_ident = _get_first(general_map, ["Identificacion del Representante Legal", "Identificacion Representante Legal"])

    tipo_ident = _extract_id_type(ident)
    ident_clean = _clean_id(ident)

    # Prioridad: identificacion del representante legal capturada por rotulo (mas estable en SECOP)
    rep_ident_final = _clean_id(rep_id_raw or rep_ident)

    # BPIM (si esta en cualquier mapa)
    bpim = _get_first(contrato_map, ["BPIM", "BPIN", "Codigo BPIM", "Codigo BPIM"])
    if not bpim:
        bpim = _get_first(general_map, ["BPIM", "BPIN", "Codigo BPIM", "Codigo BPIM"])

    bpim = _clean_bpim(bpim)

    tipo_gasto = _get_first(general_map, ["Tipo de Gasto", "Tipo Gasto"])
    if not tipo_gasto:
        tipo_gasto = _get_first(contrato_map, ["Tipo de Gasto", "Tipo Gasto"])
    tipo_proc = _determine_tipo_proceso(tipo_gasto)

    record = {
        "Numero de proceso (informativo)": num_proceso_info,
        "Numero de constancia": constancia_ok,
        "Tipo de Gasto": tipo_proc,
        "Estado del proceso": estado_proc,
        "Modalidad de contratacion": modalidad,
        "Fuente de financiacion": fuente_fin,
        "Registro Presupuestal (RP)": rp_code,
        "Certificado de disponibilidad presupuestal": cdp,
        "Numero de contrato": num_contrato,
        "Objeto del contrato": objeto,
        "Valor del contrato (COP)": valor_num,
        "Plazo de ejecucion": plazo,
        "Fecha de inicio": fecha_inicio,
        "Fecha de terminacion": fecha_fin,
        "Razon social del proponente/contratista": razon_social,
        "Tipo de identificacion": tipo_ident,
        "Identificacion del proponente/contratista": ident_clean,
        "Representante legal": rep_legal,
        "Identificación del representante legal": rep_ident_final,
        "Codigo BPIM": bpim,
        "Fuente del documento": "SECOP I (detalleProceso)",
    }

    estado_val, obs_val = _estado_validacion(record)
    # Observaciones extendidas: anadir faltantes de campos importantes de contrato/presupuesto
    obs_parts = []
    if obs_val:
        obs_parts.append(obs_val)
    # Faltantes extra
    extra_keys = ["Numero de contrato", "Plazo de ejecucion", "Fecha de inicio"]
    missing_extra = [k for k in extra_keys if not (record.get(k) or "").strip()]
    if missing_extra:
        obs_parts.append("Campos sin dato: " + ", ".join(missing_extra))
    record["Estado de validacion"] = estado_val
    record["Observaciones"] = " | ".join(obs_parts).strip(" |")

    return record


def extract_to_excel(
    constancia: str,
    out_dir: Path,
    headless: bool = False,
    template_path: Optional[Path] = None,
) -> Path:
    """
    Extrae datos del detalle SECOP I y llena la plantilla estandar (v1.2.3+).
    Genera un XLSX por constancia (la UI puede agruparlos en ZIP).
    """
    constancia_ok = validate_constancia(constancia)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if template_path is None:
        template_path = Path(__file__).parent / "templates" / "Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx"

    html = fetch_detail_html(constancia_ok, headless=headless)
    soup = BeautifulSoup(html, "html.parser")
    record = _build_record_from_soup(soup, constancia_ok)

    # Escribir en plantilla
    wb = _load_template(template_path)
    ws = wb["Resultados_Extraccion"]
    headers = [c.value for c in ws[1]]

    row_idx = _find_next_row(ws)
    _write_record_row(ws, wb, headers, record, constancia_ok, row_idx)

    out_path = out_dir / f"Resultados_Extraccion_{constancia_ok}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    return out_path


def extract_batch_to_excel(
    constancias: List[str],
    out_dir: Path,
    headless: bool = False,
    template_path: Optional[Path] = None,
    delay_seconds: float = 30.0,
    backoff_max_seconds: float = 600.0,
) -> Tuple[Path, List[Tuple[str, str]]]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if template_path is None:
        template_path = Path(__file__).parent / "templates" / "Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx"

    wb = _load_template(template_path)
    ws = wb["Resultados_Extraccion"]
    headers = [c.value for c in ws[1]]
    row_idx = _find_next_row(ws)

    errors: List[Tuple[str, str]] = []
    blocked = False
    backoff = delay_seconds

    total_constancias = len(constancias)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1280, "height": 720})
        page = context.new_page()
        try:
            if total_constancias > 2:
                warmup = random.uniform(15.0, 30.0)
                time.sleep(warmup)
            for idx, c in enumerate(constancias):
                try:
                    constancia_ok = validate_constancia(c)
                    # Pausa antes de abrir el detalle para evitar bloqueos
                    if total_constancias > 2 and idx > 0:
                        jitter = random.uniform(0.8, 1.2)
                        time.sleep(backoff * jitter)
                    html = _fetch_detail_html_with_page(page, constancia_ok)
                    soup = BeautifulSoup(html, "html.parser")
                    record = _build_record_from_soup(soup, constancia_ok)
                    _write_record_row(ws, wb, headers, record, constancia_ok, row_idx)
                    row_idx += 1
                    backoff = delay_seconds
                except SecopExtractionError as e:
                    msg = str(e)
                    errors.append((c, msg))
                    if "bloqueado" in msg.lower() or "blocked" in msg.lower():
                        blocked = True
                        break
                    backoff = min(backoff * 2, backoff_max_seconds)
                except Exception as e:
                    errors.append((c, str(e)))
                    backoff = min(backoff * 2, backoff_max_seconds)

        finally:
            context.close()
            browser.close()

    if errors:
        if "Errores" in wb.sheetnames:
            del wb["Errores"]
        _append_errors_sheet(wb, errors)

    out_path = out_dir / f"Resultados_Extraccion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    if blocked:
        errors.append(("_BLOQUEO_", "Lote detenido por bloqueo anti-DDoS. Reintenta mas tarde."))
    return out_path, errors


def append_batch_to_excel(
    constancias: List[str],
    out_path: Path,
    headless: bool = False,
    template_path: Optional[Path] = None,
    delay_seconds: float = 30.0,
    backoff_max_seconds: float = 600.0,
) -> Tuple[Path, List[Tuple[str, str]], int]:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if template_path is None:
        template_path = Path(__file__).parent / "templates" / "Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx"

    if out_path.exists():
        wb = openpyxl.load_workbook(out_path)
        if "Resultados_Extraccion" not in wb.sheetnames:
            raise SecopExtractionError("La plantilla no contiene la hoja 'Resultados_Extraccion'.")
    else:
        wb = _load_template(template_path)

    ws = wb["Resultados_Extraccion"]
    headers = [c.value for c in ws[1]]
    row_idx = _find_next_row(ws)

    errors: List[Tuple[str, str]] = []
    blocked = False
    backoff = delay_seconds
    ok_count = 0

    total_constancias = len(constancias)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1280, "height": 720})
        page = context.new_page()
        try:
            if total_constancias > 2:
                warmup = random.uniform(15.0, 30.0)
                time.sleep(warmup)
            for idx, c in enumerate(constancias):
                try:
                    constancia_ok = validate_constancia(c)
                    if total_constancias > 2 and idx > 0:
                        jitter = random.uniform(0.8, 1.2)
                        time.sleep(backoff * jitter)
                    html = _fetch_detail_html_with_page(page, constancia_ok)
                    soup = BeautifulSoup(html, "html.parser")
                    record = _build_record_from_soup(soup, constancia_ok)
                    _write_record_row(ws, wb, headers, record, constancia_ok, row_idx)
                    row_idx += 1
                    ok_count += 1
                    backoff = delay_seconds
                except SecopExtractionError as e:
                    msg = str(e)
                    errors.append((c, msg))
                    if "bloqueado" in msg.lower() or "blocked" in msg.lower():
                        blocked = True
                        break
                    backoff = min(backoff * 2, backoff_max_seconds)
                except Exception as e:
                    errors.append((c, str(e)))
                    backoff = min(backoff * 2, backoff_max_seconds)
        finally:
            context.close()
            browser.close()

    if errors:
        _append_errors_sheet(wb, errors)

    wb.save(out_path)
    if blocked:
        errors.append(("_BLOQUEO_", "Lote detenido por bloqueo anti-DDoS. Reintenta mas tarde."))
    return out_path, errors, ok_count


def extract_record_from_html(html: str, constancia_ok: str = "") -> dict:
    """Extrae un subconjunto de campos criticos desde HTML (modo OFFLINE).

    - No usa Playwright
    - No escribe Excel
    - Esta disenado para validacion y regresion de extraccion (RP y CDP)
    """
    soup = BeautifulSoup(html, "html.parser")

    rp_code = _extract_crp_code(soup)
    cdp = _extract_cdp(soup)

    return {
        "Numero de constancia": constancia_ok,
        "Registro Presupuestal (RP)": rp_code,
        "Certificado de disponibilidad presupuestal": cdp,
    }


# Compatibilidad con tu UI: permite secop_extract.main(url) o main(constancia)
def main(arg: str):
    """
    Soporta:
    - main("25-11-14555665")
    - main("https://www.contratos.gov.co/consultas/detalleProceso.do?numConstancia=25-11-14555665")
    """
    s = (arg or "").strip()
    if "numConstancia=" in s:
        const = s.split("numConstancia=", 1)[1].split("&", 1)[0]
    else:
        const = s
    return extract_to_excel(const, Path.home() / "secop_exports", headless=False)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        raise SystemExit("Uso: python secop_extract.py <numConstancia>")
    print(main(sys.argv[1]))
