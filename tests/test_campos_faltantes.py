#!/usr/bin/env python3
"""
test_campos_faltantes.py

Valida que:
1. Los nombres de columnas en secop_extract.py coincidan con la plantilla Excel
2. Todas las variables necesarias esten disponibles para los campos de identificacion
"""

import sys
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT_DIR / "scripts"
if SCRIPTS_DIR.exists():
    sys.path.insert(0, str(SCRIPTS_DIR))

# Configuracion
SECOP_EXTRACT_FILE = SCRIPTS_DIR / "secop_extract.py"
TEMPLATE_FILE = ROOT_DIR / "templates" / "Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx"

# Columnas esperadas en la plantilla (solo campos de identificacion relevantes)
EXPECTED_COLUMNS = [
    "Tipo de identificación",                        # Tipo (CC, NIT, CE)
    "Identificación del proponente/contratista",     # Numero limpio
    "Representante legal",                           # Nombre
    "Identificación del representante legal",        # Numero representante legal
]

# Variables que deben existir en secop_extract.py
REQUIRED_VARS = [
    "ident",          # raw contractor ID
    "tipo_ident",     # tipo de identificacion
    "ident_clean",    # cleaned contractor ID
    "rep_legal",      # representative name
    "rep_ident_final",  # representante legal (numero)
]

# Claves esperadas en el diccionario record
EXPECTED_RECORD_KEYS = {
    "Tipo de identificación": "tipo_ident",
    "Identificación del proponente/contratista": "ident_clean",
    "Representante legal": "rep_legal",
    "Identificación del representante legal": "rep_ident_final",
}


def test_excel_template():
    """Valida que la plantilla tiene los encabezados esperados."""
    print("[TEST 1] Validando plantilla Excel...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(TEMPLATE_FILE))
        ws = wb["Resultados_Extraccion"]
        headers = [c.value for c in ws[1] if c.value]

        missing = []
        for col in EXPECTED_COLUMNS:
            if col not in headers:
                missing.append(col)

        if missing:
            print(f"  ? FALTA: {missing}")
            return False
        print(f"  V Todas las columnas esperadas existen ({len(EXPECTED_COLUMNS)})")
        return True
    except Exception as e:
        print(f"  ? Error: {e}")
        return False


def test_secop_extract_syntax():
    """Valida que secop_extract.py tiene sintaxis correcta."""
    print("[TEST 2] Validando sintaxis de secop_extract.py...")
    try:
        import py_compile
        py_compile.compile(str(SECOP_EXTRACT_FILE), doraise=True)
        print("  V Sintaxis correcta")
        return True
    except Exception as e:
        print(f"  ? Error de sintaxis: {e}")
        return False


def test_record_dict_keys():
    """Valida que el diccionario record tiene las claves correctas."""
    print("[TEST 3] Validando claves en diccionario record...")
    try:
        with open(SECOP_EXTRACT_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        record_start = content.find('record = {')
        record_end = content.find('}', record_start) + 1
        record_block = content[record_start:record_end]

        missing_keys = []
        for key in EXPECTED_RECORD_KEYS.keys():
            if f'"{key}"' not in record_block:
                missing_keys.append(key)

        if missing_keys:
            print(f"  ? FALTA claves en record: {missing_keys}")
            return False

        print(f"  V Todas las claves del record estan presentes ({len(EXPECTED_RECORD_KEYS)})")
        return True
    except Exception as e:
        print(f"  ? Error: {e}")
        return False


def test_variables_exist():
    """Valida que todas las variables necesarias existen en secop_extract.py."""
    print("[TEST 4] Validando variables requeridas...")
    try:
        with open(SECOP_EXTRACT_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        missing_vars = []
        for var in REQUIRED_VARS:
            if f"{var} =" not in content and f"{var}=" not in content:
                missing_vars.append(var)

        if missing_vars:
            print(f"  ? FALTAN variables: {missing_vars}")
            return False

        print(f"  V Todas las variables existen ({len(REQUIRED_VARS)})")
        return True
    except Exception as e:
        print(f"  ? Error: {e}")
        return False


def test_record_value_mapping():
    """Valida que cada clave record esta mapeada a la variable correcta."""
    print("[TEST 5] Validando mapeo record variables...")
    try:
        with open(SECOP_EXTRACT_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        record_start = content.find('record = {')
        record_end = content.find('}', record_start) + 1
        record_block = content[record_start:record_end]

        errors = []
        for key, expected_var in EXPECTED_RECORD_KEYS.items():
            pattern = f'"{key}"'
            if pattern in record_block:
                start = record_block.find(pattern)
                line_end = record_block.find('
', start)
                if line_end == -1:
                    line_end = record_block.find(',', start)
                line = record_block[start:line_end]

                if expected_var not in line:
                    errors.append(f"{key} -> {expected_var} (pero la linea es: {line})")

        if errors:
            print("  ? Mapeos incorrectos:")
            for err in errors:
                print(f"      - {err}")
            return False

        print(f"  V Todos los mapeos son correctos ({len(EXPECTED_RECORD_KEYS)})")
        return True
    except Exception as e:
        print(f"  ? Error: {e}")
        return False


def main():
    print("=" * 70)
    print("VALIDACION: Campos Faltantes en Salida Excel")
    print("=" * 70)
    print()

    results = []
    results.append(("Excel Template", test_excel_template()))
    results.append(("Sintaxis secop_extract.py", test_secop_extract_syntax()))
    results.append(("Claves en record", test_record_dict_keys()))
    results.append(("Variables requeridas", test_variables_exist()))
    results.append(("Mapeo record variables", test_record_value_mapping()))

    print()
    print("=" * 70)
    print("RESUMEN")
    print("=" * 70)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for name, result in results:
        status = "V PASS" if result else "? FAIL"
        print(f"{status}: {name}")

    print()
    print(f"Resultado: {passed}/{total} pruebas pasadas")
    print()

    if passed == total:
        print("V Todos los campos estan correctamente mapeados.")
        print("  La correccion deberia resolver el problema de campos faltantes.")
        return 0
    else:
        print("? Hay problemas que requieren atencion.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
