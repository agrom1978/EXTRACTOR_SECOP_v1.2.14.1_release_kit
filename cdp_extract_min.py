from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional

import openpyxl
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


def _norm_text(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _pick_numeric_token(s: str) -> str:
    tokens = re.findall(r"\d+", s or "")
    if not tokens:
        return ""
    for t in tokens:
        if len(t) == 10:  # YYMMDD####
            return t
    for t in tokens:
        if 4 <= len(t) <= 9:
            return t
    return max(tokens, key=len)


def extract_cdp_from_html(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    # Buscar la seccion "Respaldos Presupuestales Asociados al Proceso"
    for td in soup.find_all("td"):
        if "tttablas" not in (td.get("class") or []):
            continue
        if "respaldos presupuestales asociados al proceso" not in _norm_text(td.get_text()):
            continue
        table = td.find_parent("tr").find_next("table")
        if not table:
            continue
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        header_cells = rows[0].find_all(["th", "td"])
        header = [_norm_text(c.get_text()) for c in header_cells]
        try:
            idx_num = next(i for i, h in enumerate(header) if "numero" in h and "respaldo" in h)
        except StopIteration:
            idx_num = None
        if idx_num is None:
            continue
        candidates: List[str] = []
        for r in rows[1:]:
            cells = r.find_all(["th", "td"])
            if idx_num >= len(cells):
                continue
            token = _pick_numeric_token(cells[idx_num].get_text(" ", strip=True))
            if token:
                candidates.append(token)
        if candidates:
            for c in candidates:
                if len(c) == 10:
                    return c
            return max(candidates, key=len)

    # Fallback: buscar cualquier numero en el texto completo
    text = soup.get_text(" ", strip=True)
    return _pick_numeric_token(text)


def _fetch_detail_html(constancia: str, headless: bool = False, timeout_ms: int = 120_000) -> str:
    url = f"https://www.contratos.gov.co/consultas/detalleProceso.do?numConstancia={constancia}"
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1280, "height": 720})
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            page.wait_for_timeout(1500)
            try:
                page.wait_for_selector("td.tttablas", timeout=20_000)
            except PWTimeoutError:
                pass
            page.wait_for_timeout(1200)
            return page.content()
        finally:
            context.close()
            browser.close()


def _find_next_row(ws) -> int:
    headers = [c.value for c in ws[1]]
    header_map = {_norm_text(str(h or "")): i + 1 for i, h in enumerate(headers)}
    col_a = header_map.get(_norm_text("Numero de proceso (informativo)"), 1)
    col_b = header_map.get(_norm_text("Numero de constancia"), 2)
    for r in range(2, ws.max_row + 2):
        a = ws.cell(row=r, column=col_a).value
        b = ws.cell(row=r, column=col_b).value
        if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == ""):
            return r
    return ws.max_row + 1


def _find_header_col(ws, header_name: str) -> Optional[int]:
    target = _norm_text(header_name)
    for idx, cell in enumerate(ws[1], start=1):
        if _norm_text(str(cell.value or "")) == target:
            return idx
    return None


def _write_cdp_to_template(template_path: Path, out_path: Path, cdp: str, constancia: str = "") -> Path:
    wb = openpyxl.load_workbook(template_path)
    if "Resultados_Extraccion" not in wb.sheetnames:
        raise SystemExit("La plantilla no contiene la hoja 'Resultados_Extraccion'.")
    ws = wb["Resultados_Extraccion"]
    row_idx = _find_next_row(ws)

    col_cdp = _find_header_col(ws, "Certificado de disponibilidad presupuestal")
    if col_cdp is None:
        raise SystemExit("No se encontro la columna 'Certificado de disponibilidad presupuestal' en la plantilla.")
    ws.cell(row=row_idx, column=col_cdp, value=cdp)

    if constancia:
        col_const = _find_header_col(ws, "Numero de constancia")
        if col_const is not None:
            ws.cell(row=row_idx, column=col_const, value=constancia)

    wb.save(out_path)
    return out_path


def main(
    constancia: Optional[str] = None,
    html_path: Optional[str] = None,
    template_path: Optional[str] = None,
    out_dir: Optional[str] = None,
    headless: bool = False,
) -> str:
    default_html = r"C:\Users\USUARIO\Downloads\FireShot\Detalle del proceso_ LIC 002-25.html"
    if constancia:
        html = _fetch_detail_html(constancia, headless=headless)
    else:
        html_source = html_path or default_html
        html = Path(html_source).read_text(encoding="utf-8", errors="ignore")

    cdp = extract_cdp_from_html(html)
    if not cdp:
        raise SystemExit("No se pudo extraer el CDP.")

    if template_path:
        out_dir_path = Path(out_dir or ".")
        out_dir_path.mkdir(parents=True, exist_ok=True)
        out_path = out_dir_path / f"CDP_{constancia or 'HTML'}_solo.xlsx"
        _write_cdp_to_template(Path(template_path), out_path, cdp, constancia or "")
        return str(out_path)

    return cdp


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        raise SystemExit(
            "Uso:\n"
            "  python cdp_extract_min.py --constancia 25-15-14581710 --template templates\\Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx\n"
            "  python cdp_extract_min.py --html detalle.html --template templates\\Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx\n"
        )
    args = sys.argv[1:]
    constancia = None
    html_path = None
    template_path = None
    out_dir = None
    headless = False
    i = 0
    while i < len(args):
        if args[i] == "--constancia" and i + 1 < len(args):
            constancia = args[i + 1]
            i += 2
        elif args[i] == "--html" and i + 1 < len(args):
            html_path = args[i + 1]
            i += 2
        elif args[i] == "--template" and i + 1 < len(args):
            template_path = args[i + 1]
            i += 2
        elif args[i] == "--out-dir" and i + 1 < len(args):
            out_dir = args[i + 1]
            i += 2
        elif args[i] == "--headless":
            headless = True
            i += 1
        else:
            raise SystemExit(f"Argumento no reconocido: {args[i]}")

    print(main(constancia, html_path, template_path, out_dir, headless=headless))
