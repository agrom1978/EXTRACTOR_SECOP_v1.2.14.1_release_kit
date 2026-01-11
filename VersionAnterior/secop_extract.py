# secop_extract.py
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import openpyxl
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


# -----------------------------
# Configuración
# -----------------------------
SECOP_BASE_URL = "https://www.contratos.gov.co/consultas/detalleProceso.do?numConstancia="

# Constancia tipo: 25-1-241304, 25-15-14542595, etc.
# Nota: en la práctica SECOP I usa yy-m-xxxxxx o yy-mm-xxxxxx; toleramos 1-2 dígitos en el bloque "xx".
CONSTANCIA_RE = re.compile(r"^(?P<yy>\d{2})-(?P<xx>\d{1,2})-(?P<num>\d{4,12})$")

# Soporta guiones Unicode comunes (en dash, em dash, etc.)
DASHES_RE = re.compile(r"[\u2010\u2011\u2012\u2013\u2014\u2212]")


# -----------------------------
# Errores
# -----------------------------
class SecopExtractionError(Exception):
    pass


def normalize_constancia(constancia: str) -> str:
    s = (constancia or "").strip()
    s = DASHES_RE.sub("-", s)
    s = re.sub(r"\s+", "", s)
    return s


def validate_constancia(constancia: str) -> str:
    c = normalize_constancia(constancia)
    if not CONSTANCIA_RE.match(c):
        raise SecopExtractionError(
            f"Constancia inválida: '{constancia}'. Formato esperado: YY-XX-NNNN (ej.: 25-1-241304, 25-15-14542595)"
        )
    return c


def build_url(constancia: str) -> str:
    return f"{SECOP_BASE_URL}{constancia}"


def _safe_text(node: Any) -> str:
    if node is None:
        return ""
    return node.get_text(" ", strip=True) if hasattr(node, "get_text") else str(node).strip()


def _norm_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    s = s.replace(":", "")
    return s


def _extract_kv_from_table(table) -> List[Tuple[str, str]]:
    """
    Extrae pares (label, value) de una tabla típica de 2 columnas.
    Respeta textareas (objetos largos).
    """
    pairs: List[Tuple[str, str]] = []
    if not table:
        return pairs
    for tr in table.find_all("tr"):
        tds = tr.find_all("td", recursive=False)
        if len(tds) < 2:
            continue
        label = _safe_text(tds[0]).strip()
        value_td = tds[1]
        if not label:
            continue
        textarea = value_td.find("textarea")
        if textarea:
            value = textarea.get_text("\n", strip=True)
        else:
            value = _safe_text(value_td)
        if value is None:
            value = ""
        pairs.append((label, value))
    return pairs


def _find_section_header_td(soup: BeautifulSoup, header_text: str):
    target = _norm_key(header_text)
    for td in soup.find_all("td"):
        if "tttablas" in (td.get("class") or []):
            key = _norm_key(_safe_text(td))
            if key == target or (target and target in key) or (key and key in target):
                return td
    return None


def _parse_section_kv(soup: BeautifulSoup, header_text: str) -> List[Tuple[str, str]]:
    """
    Dado un encabezado de sección (td.tttablas), encuentra la primera tabla útil posterior y la parsea como KV.
    """
    header_td = _find_section_header_td(soup, header_text)
    if not header_td:
        return []
    header_tr = header_td.find_parent("tr")
    if not header_tr:
        return []
    # Buscar tablas siguientes; tomar la primera que tenga al menos 2 filas KV plausibles
    for table in header_tr.find_all_next("table", limit=12):
        pairs = _extract_kv_from_table(table)
        # Heurística: al menos 3 pares y labels con algo de contenido
        if len(pairs) >= 3:
            return pairs
    return []



def _is_nonempty(v: str) -> bool:
    return bool(v and str(v).strip() and str(v).strip().lower() not in {"nan", "none", "null", "-"})


def _merge_maps_keep_first(*maps: dict) -> dict:
    """Merge maps preserving the first non-empty value for each key."""
    out = {}
    for mp in maps:
        if not mp:
            continue
        for k, v in mp.items():
            if k not in out or not _is_nonempty(out.get(k, "")):
                if _is_nonempty(v):
                    out[k] = v
    return out



def _digits_only(s: str) -> str:
    if not s:
        return ""
    return re.sub(r"\D+", "", str(s))

def _norm_text(s: str) -> str:
    """Normaliza texto para comparaciones tolerantes (sin tildes, sin puntuación, espacios colapsados)."""
    if s is None:
        return ""
    s = str(s)
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_row_value_by_label(soup, label_substr: str) -> str:
    """Busca valor (celda derecha) para una fila cuyo rótulo (celda izquierda) coincide de forma tolerante."""
    target = _norm_text(label_substr)
    if not target:
        return ""
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            left = _safe_text(cells[0])
            if not left:
                continue
            if target in _norm_text(left):
                return _safe_text(cells[1]).strip()
    return ""

def _find_rp_code(soup) -> str:
    """Extrae el código RP/CRP desde la tabla con encabezados 'Código|Fecha|Valor' de forma tolerante.
    No depende del título de sección (SECOP varía el encabezado).
    """
    # 1) Tabla por estructura (encabezados)
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        header_cells = rows[0].find_all(["th", "td"])
        headers = [_norm_text(_safe_text(c)) for c in header_cells]
        if not headers:
            continue
        if "codigo" in headers and "valor" in headers:
            code_idx = headers.index("codigo")
            for r in rows[1:]:
                data_cells = r.find_all(["th", "td"])
                if len(data_cells) <= code_idx:
                    continue
                raw = _safe_text(data_cells[code_idx])
                raw_digits = _digits_only(raw)
                if raw_digits and len(raw_digits) >= 6:
                    return raw_digits
    # 2) Fallback regex en texto completo
    text = soup.get_text(" ", strip=True)
    for pat in (
        r"\bRP\b\s*(?:No\.|Nro\.|#|:)?\s*(\d{6,})",
        r"\bCRP\b\s*(?:No\.|Nro\.|#|:)?\s*(\d{6,})",
    ):
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return m.group(1)
    return ""

def _parse_all_kv(soup):
    """Parsea pares etiqueta/valor de manera tolerante recorriendo toda la página.

    - Filas con 2 celdas (th/td o td/td)
    - Ignora tablas de documentos/hitos cuando no tienen estructura KV
    """
    pairs = []
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"], recursive=False)
            if len(cells) != 2:
                # algunos layouts tienen td anidados; intentar modo recursivo si no hay celdas directas
                cells = tr.find_all(["th", "td"])[:2]
            if len(cells) != 2:
                continue
            k = _safe_text(cells[0])
            v = _safe_text(cells[1])
            if _is_nonempty(k) and _is_nonempty(v):
                pairs.append((k, v))
    return pairs


def _parse_section_table(soup: BeautifulSoup, header_text: str) -> Optional[List[List[str]]]:
    """
    Parsea una tabla con encabezados (th/td) posterior a un header.
    Retorna matriz (filas) incluyendo encabezado como primera fila.
    """
    header_td = _find_section_header_td(soup, header_text)
    if not header_td:
        return None
    header_tr = header_td.find_parent("tr")
    if not header_tr:
        return None
    for table in header_tr.find_all_next("table", limit=15):
        rows = []
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"], recursive=False)
            if not cells:
                continue
            rows.append([_safe_text(c) for c in cells])
        # Heurística: tabla con encabezado y 1+ filas
        if len(rows) >= 2 and len(rows[0]) >= 2:
            return rows
    return None


def _kv_to_map(pairs: List[Tuple[str, str]]) -> Dict[str, str]:
    m: Dict[str, str] = {}
    for k, v in pairs:
        nk = _norm_key(k)
        if nk and nk not in m:  # primera ocurrencia gana
            m[nk] = (v or "").strip()
    return m


def _get_first(m: Dict[str, str], keys: List[str]) -> str:
    for k in keys:
        nk = _norm_key(k)
        # match exact
        if nk in m and m[nk]:
            return m[nk]
    # fallback: contains tokens
    for k in keys:
        nk = _norm_key(k)
        for mk, mv in m.items():
            if mv and nk in mk:
                return mv
    return ""


def _parse_numero_proceso_informativo(soup: BeautifulSoup) -> str:
    txt = soup.get_text("\n", strip=True)
    # Ej: "Detalle del Proceso Número: LIC 001-26"
    m = re.search(r"Detalle del Proceso N[úu]mero:\s*([^\n\r]+)", txt, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""


def _parse_fuente_financiacion(soup: BeautifulSoup) -> str:
    # Preferir tabla "Fuentes de Financiación"
    rows = _parse_section_table(soup, "Fuentes de Financiación")
    if rows and len(rows) >= 2:
        header = [_norm_key(h) for h in rows[0]]
        # buscamos columna "fuente"
        try:
            idx_fuente = header.index("fuente")
        except ValueError:
            idx_fuente = None
        fuentes = []
        for r in rows[1:]:
            if idx_fuente is not None and idx_fuente < len(r):
                val = (r[idx_fuente] or "").strip()
                if val:
                    fuentes.append(val)
        if fuentes:
            # dedupe manteniendo orden
            out=[]
            for f in fuentes:
                if f not in out:
                    out.append(f)
            return "; ".join(out)
    return ""


def _parse_rp_table(soup: BeautifulSoup) -> Dict[str, str]:
    """
    Extrae Código RP/CRP, Fecha y Valor desde "Registro Presupuestal del Compromiso (RP)".
    Retorna dict con claves: codigo_rp, fecha_rp, valor_rp
    """
    out = {"codigo_rp": "", "fecha_rp": "", "valor_rp": ""}
    # intentar varios encabezados posibles (SECOP I varía)
    rows = None
    for h in [
        "Registro Presupuestal del Compromiso (RP)",
        "Registro Presupuestal del Compromiso",
        "Registro Presupuestal (RP)",
        "Registro Presupuestal",
        "Registro Presupuestal del Compromiso - RP",
    ]:
        rows = _parse_section_table(soup, h)
        if rows:
            break
    if not rows:
        return out
    header = [_norm_key(h) for h in rows[0]]
    # índices posibles
    def idx_of(*cands):
        for c in cands:
            nc=_norm_key(c)
            if nc in header:
                return header.index(nc)
        return None
    i_codigo = idx_of("Código", "Codigo")
    i_fecha = idx_of("Fecha")
    i_valor = idx_of("Valor")
    # Tomar primera fila de datos válida
    for r in rows[1:]:
        if not r or len(r) < 2:
            continue
        if i_codigo is not None and i_codigo < len(r) and not out["codigo_rp"]:
            out["codigo_rp"] = (r[i_codigo] or "").strip()
        if i_fecha is not None and i_fecha < len(r) and not out["fecha_rp"]:
            out["fecha_rp"] = (r[i_fecha] or "").strip()
        if i_valor is not None and i_valor < len(r) and not out["valor_rp"]:
            out["valor_rp"] = (r[i_valor] or "").strip()
        if out["codigo_rp"] or out["fecha_rp"] or out["valor_rp"]:
            break
    return out


def fetch_detail_html(constancia: str, headless: bool = False, timeout_ms: int = 120_000) -> str:
    """
    Abre el detalle SECOP I en Playwright (visible por defecto para resolver reCAPTCHA) y retorna el HTML renderizado.
    """
    url = build_url(constancia)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1280, "height": 720})
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            # Espera razonable a tablas principales. Si hay reCAPTCHA, el usuario debe resolverlo y luego continuar.
            # Reintentamos esperar por un elemento común en el detalle.
            page.wait_for_timeout(1500)
            try:
                page.wait_for_selector("text=Información General del Proceso", timeout=20_000)
            except PWTimeoutError:
                # No forzamos fallo inmediato; puede haber variaciones o el usuario aún resolviendo captcha.
                pass
            # Espera adicional corta para render
            page.wait_for_timeout(1200)
            html = page.content()
        finally:
            context.close()
            browser.close()
    return html



def _extract_digits(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    d = re.sub(r"[^0-9]", "", s)
    return d

def _clean_id(s: str) -> str:
    """Identificación limpia (solo dígitos)."""
    return _extract_digits(s)


def _clean_bpim(raw: str) -> str:
    """Normaliza BPIN/BPIM a un único código numérico sin leyendas.

    Objetivo: YYYY + consecutivo (p.ej. 20250000003856). El portal puede traer:
    - 'BPIM: 2025 00000003856'
    - 'Código BPIM Año 2025 20250000003856'
    - Variaciones con puntos/espacios.
    """
    s = (raw or "").strip()
    if not s:
        return ""
    # extraer todos los grupos numéricos
    nums = re.findall(r"\d+", s)
    if not nums:
        return ""

    # Si ya existe un token largo que empieza por 20xx, usar el más largo.
    candidates = [n for n in nums if len(n) >= 12 and n.startswith("20")]
    if candidates:
        # elegir el más largo; si empate, el primero
        candidates.sort(key=lambda x: (-len(x), nums.index(x)))
        return candidates[0]

    # Caso típico separado: '2025' + '00000003856' -> concatenar
    year_tokens = [n for n in nums if len(n) == 4 and n.startswith("20")]
    if year_tokens:
        year = year_tokens[0]
        # elegir el token más largo que NO sea el año y que no empiece por 20
        others = [n for n in nums if n != year]
        if others:
            other = max(others, key=len)
            merged = year + other
            if len(merged) >= 12:
                return merged

    # Fallback: si hay un token largo sin '20', devolver el más largo
    other_long = [n for n in nums if len(n) >= 10]
    if other_long:
        return max(other_long, key=len)

    return ""


def _clean_fuente_financiacion(raw: str) -> str:
    """Normaliza la fuente de financiación a un catálogo corto y útil.

    Ejemplos de salida:
    - 'Sistema General de Regalías'
    - 'Sistema General de Participaciones (SGP)'
    - 'Recursos propios'
    - 'Otros recursos'
    """
    s = (raw or "").strip()
    if not s:
        return ""
    n = _norm_key(s)

    # Prioridad por especificidad
    if "regalia" in n or "sgr" in n:
        return "Sistema General de Regalías"
    if "participacion" in n or re.search(r"\bsgp\b", n):
        return "Sistema General de Participaciones (SGP)"
    if "propio" in n:
        return "Recursos propios"
    if "otro" in n or "otros" in n:
        return "Otros recursos"

    # Si no clasifica, devolver el texto original recortado (sin encabezados comunes)
    # Quitar palabras de encabezado frecuentes
    for junk in ["fuente", "financiacion", "financiación", "valor", "tipo"]:
        n = n.replace(junk, "").strip()
    # Si quedó vacío, devolver original
    return s if n else s


def _clean_money(s: str) -> str:
    """Convierte un valor monetario a entero (COP) truncando decimales de forma segura.

    Evita inflar el valor cuando la fuente trae ',00' o '.00' y se eliminan separadores.
    Soporta formatos típicos:
      - es-CO: 608.603.520.000,00
      - en-US: 608,603,520,000.00
      - enteros: 608603520000
    Devuelve string de dígitos (sin separadores)."""
    s = (s or "").strip()
    if not s:
        return ""
    # Dejar sólo dígitos y separadores de miles/decimales
    raw = re.sub(r"[^0-9,\.]", "", s)
    if not raw:
        return ""

    # Caso mixto con ambos separadores: decidir decimal por el último separador
    if "," in raw and "." in raw:
        last_comma = raw.rfind(",")
        last_dot = raw.rfind(".")
        if last_comma > last_dot:
            # decimal = ',' ; miles = '.'
            int_part = raw.split(",", 1)[0]
            return re.sub(r"[^0-9]", "", int_part)
        else:
            # decimal = '.' ; miles = ','
            int_part = raw.split(".", 1)[0]
            return re.sub(r"[^0-9]", "", int_part)

    # Sólo coma
    if "," in raw and "." not in raw:
        left, right = raw.split(",", 1)
        # si son 2 dígitos al final, asumimos decimal
        if right.isdigit() and len(right) == 2:
            return re.sub(r"[^0-9]", "", left)
        # si no, es separador de miles
        return re.sub(r"[^0-9]", "", raw)

    # Sólo punto
    if "." in raw and "," not in raw:
        left, right = raw.split(".", 1)
        if right.isdigit() and len(right) == 2:
            return re.sub(r"[^0-9]", "", left)
        return re.sub(r"[^0-9]", "", raw)

    # Sólo dígitos
    return re.sub(r"[^0-9]", "", raw)

def _extract_rp_code(raw: str) -> str:
    """Normaliza RP/CRP a un código simple (solo dígitos)."""
    s = (raw or "").strip()
    if not s:
        return ""
    m = re.search(r"(\d{6,})", s)
    if m:
        return m.group(1)
    return re.sub(r"[^\d]", "", s)



def _determine_tipo_proceso(bpim: str, fuente_fin: str, rp_code: str) -> str:
    # Regla simple y conservadora: si hay BPIM -> Inversión.
    if bpim:
        return "Inversión"
    # Si hay fuente de financiación o RP, suele ser inversión en contexto municipal.
    if fuente_fin or rp_code:
        return "Inversión"
    return "Otro"


def _estado_validacion(record: Dict[str, str]) -> Tuple[str, str]:
    # Campos críticos para que el cuadro sea útil
    critical = {
        "Modalidad de contratación": record.get("Modalidad de contratación", ""),
        "Objeto del contrato": record.get("Objeto del contrato", ""),
        "Valor del contrato (COP)": record.get("Valor del contrato (COP)", ""),
        "Razón social del proponente/contratista": record.get("Razón social del proponente/contratista", ""),
    }

    missing = [k for k, v in critical.items() if not (v or "").strip()]
    if not missing:
        return "Completo", ""
    if len(missing) == 1:
        return "Revisión", f"Falta: {missing[0]}"
    return "Incompleto", "Faltan: " + "; ".join(missing)


def _load_template(template_path: Path):
    if not template_path.exists():
        raise SecopExtractionError(f"No se encontró la plantilla: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    if "Resultados_Extraccion" not in wb.sheetnames:
        raise SecopExtractionError("La plantilla no contiene la hoja 'Resultados_Extraccion'.")
    return wb


def _find_next_row(ws) -> int:
    """
    Encuentra la siguiente fila vacía basada en columnas A y B (A=Numero proceso info, B=Constancia),
    porque C tiene fórmula y no sirve para detectar vacío.
    """
    for r in range(2, ws.max_row + 2):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == ""):
            return r
    return ws.max_row + 1


def extract_to_excel(
    constancia: str,
    out_dir: Path,
    headless: bool = False,
    template_path: Optional[Path] = None,
) -> Path:
    """
    Extrae datos del detalle SECOP I y llena la plantilla estándar (v1.2.3+).
    Genera un XLSX por constancia (la UI puede agruparlos en ZIP).
    """
    constancia_ok = validate_constancia(constancia)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if template_path is None:
        template_path = Path(__file__).parent / "templates" / "Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx"

    html = fetch_detail_html(constancia_ok, headless=headless)
    soup = BeautifulSoup(html, "html.parser")

    # Extracciones dirigidas (sin depender de secciones): representante legal e RP
    rep_name_raw = _find_row_value_by_label(soup, "Nombre del Representante Legal")
    rep_id_raw = _find_row_value_by_label(soup, "Identificación del Representante Legal")
    rp_code = _find_rp_code(soup)

    # 0) Baseline KV tolerante (anti-regresión)
    baseline_pairs = _parse_all_kv(soup)
    baseline_map = _kv_to_map(baseline_pairs)

    # 1) General (KV por sección) — si no se encuentra, se apoya en baseline_map
    general_pairs = _parse_section_kv(soup, "Información General del Proceso")
    general_map = _merge_maps_keep_first(_kv_to_map(general_pairs), baseline_map)

    # 2) Contrato (KV por sección) — si no se encuentra, se apoya en baseline_map
    contrato_pairs = _parse_section_kv(soup, "Información del Contrato")
    contrato_map = _merge_maps_keep_first(_kv_to_map(contrato_pairs), baseline_map)

    # 3) Presupuestal (RP table + fallback KV)
    rp = _parse_rp_table(soup)


    # Campo informativo "Número de proceso"
    num_proceso_info = _parse_numero_proceso_informativo(soup)

    modalidad = _get_first(general_map, ["Tipo de Proceso", "Modalidad de Contratación", "Modalidad"])
    estado_proc = _get_first(general_map, ["Estado del Proceso", "Estado del Contrato", "Estado"])

    fuente_fin = _parse_fuente_financiacion(soup)
    if not fuente_fin:
        fuente_fin = _get_first(general_map, ["Fuente de Financiación", "Fuentes de Financiación", "Fuente"])


    fuente_fin = _clean_fuente_financiacion(fuente_fin)

    # Contrato info
    num_contrato = _get_first(contrato_map, ["Número del Contrato", "No. Contrato", "Contrato No", "Número de Contrato"])
    objeto = _get_first(contrato_map, ["Objeto del Contrato", "Objeto"])
    valor = _get_first(contrato_map, ["Cuantía Definitiva del Contrato", "Cuantía del Contrato", "Valor del Contrato", "Cuantía", "Valor"])
    valor_num = _clean_money(valor)

    plazo = _get_first(contrato_map, ["Plazo de Ejecución del Contrato", "Plazo de Ejecución", "Plazo"])
    fecha_inicio = _get_first(contrato_map, ["Fecha de Inicio de Ejecución del Contrato", "Fecha de Inicio", "Fecha inicio"])
    fecha_fin = _get_first(contrato_map, ["Fecha de Terminación del Contrato", "Fecha de Terminación", "Fecha fin", "Fecha terminacion"])

    razon_social = _get_first(contrato_map, ["Nombre o Razón Social del Contratista", "Nombre o Razon Social del Contratista", "Contratista", "Adjudicatario"])
    ident = _get_first(contrato_map, ["Identificación del Contratista", "Identificacion del Contratista", "NIT del Contratista", "NIT", "Cédula", "Cedula", "Identificación", "Identificacion"])
    if not ident:
        ident = _get_first(general_map, ["Identificación", "Identificacion", "NIT", "Cédula", "Cedula"])
    rep_legal = _get_first(contrato_map, ["Nombre del Representante Legal del Contratista", "Representante Legal", "Representante"])
    if not rep_legal:
        rep_legal = _get_first(general_map, ["Representante Legal", "Representante"])
    rep_ident = _get_first(contrato_map, ["Identificación del Representante Legal del Contratista", "Identificacion Representante Legal", "Cédula Representante", "Cedula Representante", "Identificación Representante", "Identificacion Representante"]) 

    ident_clean = _clean_id(ident)
    rep_ident_clean = _clean_id(rep_ident)

    # BPIM (si está en cualquier mapa)
    bpim = _get_first(contrato_map, ["BPIM", "BPIN", "Código BPIM", "Codigo BPIM"])
    if not bpim:
        bpim = _get_first(general_map, ["BPIM", "BPIN", "Código BPIM", "Codigo BPIM"])


    bpim = _clean_bpim(bpim)

    crp_raw = rp.get("codigo_rp") or _get_first(contrato_map, ["Código RP", "Codigo RP", "RP", "CRP", "Registro Presupuestal", "Registro Presupuestal del Compromiso (RP)"])
    if not crp_raw:
        crp_raw = _get_first(general_map, ["Código RP", "Codigo RP", "RP", "CRP", "Registro Presupuestal"])
    crp = _extract_rp_code(crp_raw)
    rp_clean = crp

    tipo_proc = _determine_tipo_proceso(bpim, fuente_fin, crp)

    record = {
        "Número de proceso (informativo)": num_proceso_info,
        "Número de constancia": constancia_ok,
        "Tipo de proceso": tipo_proc,
        "Estado del proceso": estado_proc,
        "Modalidad de contratación": modalidad,
        "Fuente de financiación": fuente_fin,
        "Código Registro Presupuestal (CRP)": crp,
        "Registro Presupuestal (RP/CRP) (limpio)": rp_clean,
        "Número de contrato": num_contrato,
        "Objeto del contrato": objeto,
        "Valor del contrato (COP)": valor_num,
        "Plazo de ejecución": plazo,
        "Fecha de inicio": fecha_inicio,
        "Fecha de terminación": fecha_fin,
        "Razón social del proponente/contratista": razon_social,
        "Identificación del proponente (CC/NIT)": ident,
        "Identificación del proponente/contratista (limpio)": ident_clean,
        "Representante legal": rep_legal,
        "Identificación representante legal": rep_ident,
        "Identificación del representante legal (limpio)": rep_ident_clean,
        "Código BPIM": bpim,
        "Fuente del documento": "SECOP I (detalleProceso)",
    }

    estado_val, obs_val = _estado_validacion(record)
    # Observaciones extendidas: añadir faltantes de campos importantes de contrato/presupuesto
    obs_parts = []
    if obs_val:
        obs_parts.append(obs_val)
    # Faltantes extra
    extra_keys = ["Número de contrato", "Plazo de ejecución", "Fecha de inicio"]  # RP/CRP puede no estar publicado; no se fuerza como faltante
    missing_extra = [k for k in extra_keys if not (record.get(k) or "").strip()]
    if missing_extra:
        obs_parts.append("Campos sin dato: " + ", ".join(missing_extra))
    record["Estado de validación"] = estado_val
    record["Observaciones"] = " | ".join(obs_parts).strip(" |")

    # Escribir en plantilla
    wb = _load_template(template_path)
    ws = wb["Resultados_Extraccion"]
    headers = [c.value for c in ws[1]]

    row_idx = _find_next_row(ws)
    # escribe columnas existentes
    for col_idx, h in enumerate(headers, start=1):
        if not h:
            continue
        if h == "Abrir detalle":
            # Hipervínculo directo (visible en más visores que la fórmula).
            cell = ws.cell(row=row_idx, column=col_idx)
            base = ""
            if "Config" in wb.sheetnames:
                base = str(wb["Config"]["B1"].value or "").strip()
            if not base:
                base = SECOP_BASE_URL
            cell.value = "Abrir"
            cell.hyperlink = base + constancia_ok
            try:
                cell.style = "Hyperlink"
            except Exception:
                pass
            continue
        if h in record:
            ws.cell(row=row_idx, column=col_idx, value=record[h])

    out_path = out_dir / f"Resultados_Extraccion_{constancia_ok}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    return out_path


# Compatibilidad con tu UI: permite secop_extract.main(url) o main(constancia)
def main(arg: str):
    """
    Soporta:
    - main("25-11-14555665")
    - main("https://www.contratos.gov.co/consultas/detalleProceso.do?numConstancia=25-11-14555665")
    """
    s = (arg or "").strip()
    if "numConstancia=" in s:
        const = s.split("numConstancia=", 1)[1].split("&", 1)[0]
    else:
        const = s
    return extract_to_excel(const, Path.home() / "secop_exports", headless=False)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        raise SystemExit("Uso: python secop_extract.py <numConstancia>")
    print(main(sys.argv[1]))
