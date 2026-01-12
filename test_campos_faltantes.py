#!/usr/bin/env python3
"""
test_campos_faltantes.py

Valida que:
1. Los nombres de columnas en secop_extract.py coincidan con la plantilla Excel
2. Todas las variables necesarias estén disponibles para los campos de identificación
"""

import sys
from pathlib import Path

# Configuración
SECOP_EXTRACT_FILE = Path(__file__).parent / "secop_extract.py"
TEMPLATE_FILE = Path(__file__).parent / "templates" / "Plantilla_Salida_EXTRACTOR_SECOP_v1.2.10.xlsx"

# Columnas esperadas en la plantilla (solo campos de identificación relevantes)
EXPECTED_COLUMNS = [
    "Identificación del proponente (CC/NIT)",           # Col 17 (raw)
    "Identificación del proponente/contratista (limpio)",  # Col 18 (limpio)
    "Representante legal",                               # Col 19
    "Identificación representante legal",                # Col 20 (raw)
    "Identificación del representante legal (limpio)",   # Col 21 (limpio)
]

# Variables que deben existir en secop_extract.py
REQUIRED_VARS = [
    "ident",                     # raw contractor ID
    "ident_clean",              # cleaned contractor ID
    "rep_legal",                # representative name
    "rep_ident_final",          # final representative ID (prioritized)
    "rep_ident_clean_final",    # cleaned representative ID
]

# Claves esperadas en el diccionario record
EXPECTED_RECORD_KEYS = {
    "Identificación del proponente (CC/NIT)": "ident",
    "Identificación del proponente/contratista (limpio)": "ident_clean",
    "Representante legal": "rep_legal",
    "Identificación representante legal": "rep_ident_final",
    "Identificación del representante legal (limpio)": "rep_ident_clean_final",
}


def test_excel_template():
    """Valida que la plantilla tiene los encabezados esperados."""
    print("[TEST 1] Validando plantilla Excel...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(TEMPLATE_FILE))
        ws = wb["Resultados_Extraccion"]
        headers = [c.value for c in ws[1] if c.value]
        
        missing = []
        for col in EXPECTED_COLUMNS:
            if col not in headers:
                missing.append(col)
        
        if missing:
            print(f"  ❌ FALTA: {missing}")
            return False
        print(f"  ✓ Todas las columnas esperadas existen ({len(EXPECTED_COLUMNS)})")
        return True
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return False


def test_secop_extract_syntax():
    """Valida que secop_extract.py tiene sintaxis correcta."""
    print("[TEST 2] Validando sintaxis de secop_extract.py...")
    try:
        import py_compile
        py_compile.compile(str(SECOP_EXTRACT_FILE), doraise=True)
        print("  ✓ Sintaxis correcta")
        return True
    except Exception as e:
        print(f"  ❌ Error de sintaxis: {e}")
        return False


def test_record_dict_keys():
    """Valida que el diccionario record tiene las claves correctas."""
    print("[TEST 3] Validando claves en diccionario record...")
    try:
        with open(SECOP_EXTRACT_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Buscar el diccionario record
        record_start = content.find('record = {')
        record_end = content.find('}', record_start) + 1
        record_block = content[record_start:record_end]
        
        missing_keys = []
        for key in EXPECTED_RECORD_KEYS.keys():
            if f'"{key}"' not in record_block:
                missing_keys.append(key)
        
        if missing_keys:
            print(f"  ❌ FALTA claves en record: {missing_keys}")
            return False
        
        print(f"  ✓ Todas las claves del record están presentes ({len(EXPECTED_RECORD_KEYS)})")
        return True
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return False


def test_variables_exist():
    """Valida que todas las variables necesarias existen en secop_extract.py."""
    print("[TEST 4] Validando variables requeridas...")
    try:
        with open(SECOP_EXTRACT_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        
        missing_vars = []
        for var in REQUIRED_VARS:
            # Buscar asignación de la variable
            if f"{var} =" not in content and f"{var}=" not in content:
                missing_vars.append(var)
        
        if missing_vars:
            print(f"  ❌ FALTAN variables: {missing_vars}")
            return False
        
        print(f"  ✓ Todas las variables existen ({len(REQUIRED_VARS)})")
        return True
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return False


def test_record_value_mapping():
    """Valida que cada clave record está mapeada a la variable correcta."""
    print("[TEST 5] Validando mapeo record → variables...")
    try:
        with open(SECOP_EXTRACT_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        
        record_start = content.find('record = {')
        record_end = content.find('}', record_start) + 1
        record_block = content[record_start:record_end]
        
        errors = []
        for key, expected_var in EXPECTED_RECORD_KEYS.items():
            # Buscar línea: "key": variable
            pattern = f'"{key}"'
            if pattern in record_block:
                # Extraer la línea
                start = record_block.find(pattern)
                line_end = record_block.find('\n', start)
                if line_end == -1:
                    line_end = record_block.find(',', start)
                line = record_block[start:line_end]
                
                if expected_var not in line:
                    errors.append(f"{key} → {expected_var} (pero la línea es: {line})")
        
        if errors:
            print(f"  ❌ Mapeos incorrectos:")
            for err in errors:
                print(f"      - {err}")
            return False
        
        print(f"  ✓ Todos los mapeos son correctos ({len(EXPECTED_RECORD_KEYS)})")
        return True
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return False


def main():
    print("=" * 70)
    print("VALIDACIÓN: Campos Faltantes en Salida Excel")
    print("=" * 70)
    print()
    
    results = []
    results.append(("Excel Template", test_excel_template()))
    results.append(("Sintaxis secop_extract.py", test_secop_extract_syntax()))
    results.append(("Claves en record", test_record_dict_keys()))
    results.append(("Variables requeridas", test_variables_exist()))
    results.append(("Mapeo record→variables", test_record_value_mapping()))
    
    print()
    print("=" * 70)
    print("RESUMEN")
    print("=" * 70)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for name, result in results:
        status = "✓ PASS" if result else "❌ FAIL"
        print(f"{status}: {name}")
    
    print()
    print(f"Resultado: {passed}/{total} pruebas pasadas")
    print()
    
    if passed == total:
        print("✓ Todos los campos están correctamente mapeados.")
        print("  La corrección debería resolver el problema de campos faltantes.")
        return 0
    else:
        print("❌ Hay problemas que requieren atención.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
